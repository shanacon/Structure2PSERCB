from time import time
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
import time
import os
import sys
from inspect import currentframe, getframeinfo
from LogSystem import *
def myexcepthook(type, value, traceback, oldhook=sys.excepthook):
    oldhook(type, value, traceback)
    input("Press RETURN. ")
sys.excepthook = myexcepthook
InitList_en = ['name', 'type', 'Bc/Dc', 'Hc', 'Los(%)', 'No1', 'Num1', 'No2', 'Num2', 'h1', 'No', 'Num', 'S', 'Nci', 'whichFloor']
InitList_zh= ['斷面名稱', '型式', '寬度/直徑', '深度', '主筋鋼筋比', '主筋號數(1)', '主筋根數(1)', '主筋號數(2)', '主筋根數(2)', 
              '一樓柱淨高', '橫向箍、繫筋號數', '橫向箍、繫筋根數', '箍筋間距', '柱根數', '所在樓層']
class Case :
    def __init__(self, name, type, BC, HC, No1, Num1, No2, Num2, H1, No, Numx, Numy, S, Nci, whichFloor):
        self.name = name  
        self.type = type  
        self.BC = BC  
        self.HC = HC  
        self.No1 = No1  
        self.Num1 = Num1  
        self.No2 = No2  
        self.Num2 = Num2  
        self.H1 = H1  
        self.No = No  
        self.Numx = Numx
        self.Numy = Numy
        self.S = S  
        self.Nci = Nci  
        self.whichFloor = whichFloor
FHdic = {}
CNdic = {}
ALLCASE = []
Num2ExpList = []
# load CXX
CXX = ReadFile('CXX.DAT', os.path.basename(__file__))
## get Floor and Cross section
CandF = CXX.readline().split()
try :
    C = int(CandF[0])
    F = int(CandF[1])
except Exception as e:
    WriteEx()
    sys.exit('CXX Read File Error.')
print('Makeing dictionary from CNK1 and MASS103...')
## load CNK
CNK1 = ReadFile('CNK1.INP', os.path.basename(__file__))
## Load useless data
KeepRead  = True
while KeepRead :
    line = CNK1.readline()
    if line.find('col.data') != -1:
        KeepRead = False
    if line == '' :
        WriteError('CNK1 Read File Error. No col.data.', os.path.basename(__file__))
        sys.exit('CNK1 Read File Error. No col.data.')
CNK1.readline()
## make dictionary
for i in range(C):
    try :
        CNdic[i] = int(CNK1.readline().split(',')[2])
    except Exception as e:
        WriteEx()
        sys.exit('CNK1 Read File Error. Line can not split normally.')
## load MASS103
MASS = ReadFile('MASS103.INP', os.path.basename(__file__))
## Load useless data
KeepRead  = True
while KeepRead :
    line = MASS.readline()
    if line.find('FLOOR') != -1:
        KeepRead = False
    if line == '' :
        WriteError('MASS103 Read File Error. No FLOOR.', os.path.basename(__file__))
        sys.exit('MASS103 Read File Error. No FLOOR.')
## get h1 from MASS
try :
    tmp = "{:.1f}".format(float(MASS.readline().split()[1]) * 100.0)
except Exception as e:
        WriteEx()
        sys.exit('MASS103 Read File Error. Line can not split normally.')
for i in range(F) :
    line = MASS.readline().split()
    FHdic[i] = tmp
    try :
        tmp = "{:.1f}".format(float(line[1]) * 100.0)
    except Exception as e:
        WriteEx()
        sys.exit('MASS103 Read File Error. Line can not split normally.')
print('Make dictionary complete')
print('Loading data from CXX...')
## Load useless data in CXX
for i in range(C):
    CXX.readline()
for i in range(F):
    CXX.readline()
for i in range(F):
    CXX.readline()
CXX_Data = CXX.readlines()
LineLen = len(CXX_Data)
Progress = 0
CaseCXX = 0
NowC = 0
NowF = 0
while CaseCXX < LineLen:
    ## set lines
    lines = []
    try :
        lines.append(CXX_Data[CaseCXX].split())
        lines.append(CXX_Data[CaseCXX + 1].split())
        lines.append(CXX_Data[CaseCXX + 2].split())
        lines.append(CXX_Data[CaseCXX + 3].split())
        lines.append(CXX_Data[CaseCXX + 4].split())
        lines.append(CXX_Data[CaseCXX + 5].split())
    except Exception as e:
        WriteEx()
        sys.exit('CXXData Out of range.')
    #floor and name
    try :
        floor = lines[0][0] + 'F'
        name = floor + lines[0][2]
    except Exception as e:
        WriteEx()
        sys.exit('CXX: line 0 Out of range. Doing floor name.')
    ## BC HC type
    try :
        BC = float(lines[1][0])
        HC = float(lines[1][1])
    except Exception as e:
        WriteEx()
        sys.exit('CXX: line 1 Out of range. Doing BC HC.')
    if BC == 0 and HC == 0:
        CaseCXX = CaseCXX + 6
        NowF = (NowF + 1) % F
        NowC = (NowC + 1) % C
        continue
    if HC == 0 or BC == 0:
        type = 'CIRL'
    else :
        type = 'RECT'
    ## No1 No2
    try :
        No1 = lines[2][0]
        if lines[2][1] == '0':
            No2 = str(int(No1) - 1)
        else :
            No2 = lines[2][1]
    except Exception as e:
        WriteEx()
        sys.exit('CXX: line 2 Out of range. Doing No1 No2.')
    No1 = '#' + No1
    No2 = '#' + No2
    ## Num1 Num2
    try :
        Num1 = (int(lines[3][0]) + int(lines[4][0])) * 2 - 4
        if Num1 < 0:
            Num1 = 0
        ## Exception of Num2
        if lines[3][1] != '0' or lines[3][2] != '0' or lines[4][1] != '0' or lines[4][2] != '0':
            Num2ExpList.append(name)
    except Exception as e:
        WriteEx()
        sys.exit('CXX: line 3 or 4 Out of range. Doing Num1.')
    Num2 = 0
    try :
        H1 = float(FHdic[NowF])
    except Exception as e:
        WriteEx()
        sys.exit('Dictionary Key Error. Doing H1.')
    try :
        if lines[5][0] != lines[5][3]:
            sys.exit('Error in No. First Element diff with last Element in line 5.')
        No = '#' + lines[5][0]
        Numx = int(lines[4][3]) + 2
        Numy = int(lines[3][3]) + 2
        S = int(lines[5][1])
    except Exception as e:
        WriteEx()
        sys.exit('CXX: line 4 or 5 Out of range. Doing No Numxy S.')
    try :
        Nci = CNdic[NowC]
    except Exception as e:
        WriteEx()
        sys.exit('Dictionary Key Error. Doing Nci.')
    whichFloor = floor
    ALLCASE.append(Case(name, type, BC, HC, No1, Num1, No2, Num2, H1, No, Numx, Numy, S, Nci, whichFloor))
    CaseCXX = CaseCXX + 6
    NowF = (NowF + 1) % F
    NowC = (NowC + 1) % C
    if float(CaseCXX / LineLen) * 10.0 > Progress:
        Progress = Progress + 1
        print("\r", end = '')
        print('[', end = '')
        for i in range(Progress):
            print('|', end = '')
        for i in range(10 - Progress):
            print(' ', end = '')
        print(']', end = '')
        time.sleep(0.05)
# initial template excel
print('Initailize excel...')
NewWb = Workbook()
sheetX = NewWb.active
sheetX.title = '一般柱-X'
NewWb.create_sheet('一般柱-Y')
sheetY = NewWb['一般柱-Y']
for i in range(65,80):
    sheetX.column_dimensions[chr(i)].width = 15.0
    sheetY.column_dimensions[chr(i)].width = 15.0
## SheetX
SheetRange = sheetX['A1':'O1']
iX = 0
for item in SheetRange[0]:
    item.value = InitList_en[iX]
    item.font = Font(name = 'Times New Roman', bold=True, size = 14)
    item.alignment = Alignment(horizontal = 'center')
    iX = iX + 1
SheetRange = sheetX['A2':'O2']
iX = 0
for item in SheetRange[0]:
    item.value = InitList_zh[iX]
    item.font = Font(name = 'Times New Roman', bold=True, size = 14)
    item.alignment = Alignment(horizontal = 'center')
    iX = iX + 1
## SheetY
SheetRange = sheetY['A1':'O1']
iY = 0
for item in SheetRange[0]:
    item.value = InitList_en[iY]
    item.font = Font(name = 'Times New Roman', bold=True, size = 14)
    item.alignment = Alignment(horizontal = 'center')
    iY = iY + 1
SheetRange = sheetY['A2':'O2']
iY = 0
for item in SheetRange[0]:
    item.value = InitList_zh[iY]
    item.font = Font(name = 'Times New Roman', bold=True, size = 14)
    item.alignment = Alignment(horizontal = 'center')
    iY = iY + 1
print('Initail excel complete')
## input data to excel
print('Writing data to sheet...')
CaseCount = 3
Caselen = len(ALLCASE)
Progress = 0
for case in ALLCASE :
    CountStr = str(CaseCount)
    ## X
    sheetX['A' + CountStr] = case.name
    sheetX['B' + CountStr] = case.type
    sheetX['C' + CountStr] = case.HC
    sheetX['D' + CountStr] = case.BC
    sheetX['F' + CountStr] = case.No1
    sheetX['G' + CountStr] = case.Num1  
    sheetX['H' + CountStr] = case.No2  
    sheetX['I' + CountStr] = case.Num2  
    sheetX['J' + CountStr] = case.H1  
    sheetX['K' + CountStr] = case.No  
    sheetX['L' + CountStr] = case.Numx
    sheetX['M' + CountStr] = case.S  
    sheetX['N' + CountStr] = case.Nci  
    sheetX['O' + CountStr] = case.whichFloor
    ## Y
    sheetY['A' + CountStr] = case.name
    sheetY['B' + CountStr] = case.type
    sheetY['C' + CountStr] = case.BC
    sheetY['D' + CountStr] = case.HC
    sheetY['F' + CountStr] = case.No1
    sheetY['G' + CountStr] = case.Num1  
    sheetY['H' + CountStr] = case.No2  
    sheetY['I' + CountStr] = case.Num2  
    sheetY['J' + CountStr] = case.H1  
    sheetY['K' + CountStr] = case.No  
    sheetY['L' + CountStr] = case.Numy
    sheetY['M' + CountStr] = case.S  
    sheetY['N' + CountStr] = case.Nci  
    sheetY['O' + CountStr] = case.whichFloor
    CaseCount = CaseCount + 1
    if float(CaseCount / Caselen) * 10.0 > Progress:
        Progress = Progress + 1
        print("\r", end = '')
        print('[', end = '')
        for i in range(Progress):
            print('|', end = '')
        for i in range(10 - Progress):
            print(' ', end = '')
        print(']', end = '')
        time.sleep(0.05)
print()
print('Complete write data to sheet')
print('Writing data to excel...\n')
## Final Adjustment
for row in range(2, sheetX.max_row):
    for column in range(sheetX.max_column):
        sheetX.cell(row=row+1,column=column+1).alignment = Alignment(horizontal = 'center')
        sheetX.cell(row=row+1,column=column+1).font = Font(name = 'Times New Roman', size = 14)
for row in range(2, sheetY.max_row):
    for column in range(sheetY.max_column):
        sheetY.cell(row=row+1,column=column+1).alignment = Alignment(horizontal = 'center')
        sheetY.cell(row=row+1,column=column+1).font = Font(name = 'Times New Roman', size = 14)
try:
    NewWb.save('test.xlsx')
except PermissionError as e:
    WriteEx()
    sys.exit('\nPermission Error. Please close the excel file and try again.')
# handle excption
for item in Num2ExpList:
    print(item + ' may has not zero in Num2.')
print('\nComplete!!')
os.system('pause')
